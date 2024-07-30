from openpyxl import Workbook
from django.shortcuts import render, redirect, get_object_or_404
from .forms import EmployeeForm, EducationForm
from django.views.generic.list import ListView
from .models import Employee, Education
from django.db.models import Q
from django.http import HttpResponse
import datetime
import pandas as pd
# from openpyxl import load_workbook
from django.urls import reverse
from django.contrib import messages


"""
def employee_create_view(request):
    if request.method == "POST":
        employee_form = EmployeeForm(request.POST)
        education_forms = [
            EducationForm(request.POST, prefix=str(i))
            for i in range(int(request.POST["education_count"]))
        ]
        if employee_form.is_valid() and all([ef.is_valid() for ef in education_forms]):
            employee = employee_form.save()
            for ef in education_forms:
                education = ef.save(commit=False)
                education.employee = employee
                education.save()
            return redirect("all_employee")
    else:
        employee_form = EmployeeForm()
        education_forms = [EducationForm(prefix="0")]
    return render(
        request,
        "employee_form.html",
        {"employee_form": employee_form, "education_forms": education_forms},
    )
"""


def employee_create_view(request):
    if request.method == "POST":
        employee_form = EmployeeForm(request.POST)
        education_forms = [
            EducationForm(request.POST, prefix=str(i))
            for i in range(int(request.POST["education_count"]))
        ]
        if employee_form.is_valid() and all([ef.is_valid() for ef in education_forms]):
            employee = employee_form.save(commit=False)
            # Save skills as a comma-separated string
            employee.skills = ", ".join(request.POST.getlist("skills"))
            employee.save()
            for ef in education_forms:
                education = ef.save(commit=False)
                education.employee = employee
                education.save()
            return redirect("all_employee")
    else:
        employee_form = EmployeeForm()
        education_forms = [EducationForm(prefix="0")]
    return render(
        request,
        "employee_form.html",
        {"employee_form": employee_form, "education_forms": education_forms},
    )


class AllEmployee(ListView):
    model = Employee
    template_name = "allEmployee.html"
    context_object_name = "employees"
    paginate_by = 2


# can be searched via email or name
class SeatchEmployee(ListView):
    model = Employee
    template_name = "searchResult.html"
    context_object_name = "employees"

    def get_queryset(self):
        query = self.request.GET.get("q")
        return Employee.objects.filter(
            Q(name__icontains=query) | Q(email__icontains=query)
        )


# for education info of each person
def getEducation(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    education_list = Education.objects.filter(employee=employee)
    return render(
        request, "edu.html", {"education_list": education_list, "employee": employee}
    )


# this is working properly as i want now
def export_employees_to_excel(request):
    # Get file name from request (or set default value)
    file_name = request.GET.get("file_name", "employees.xlsx")

    # Create a workbook and sheets
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Employees"

    # Add headers to Employees sheet
    ws1.append(["ID", "Name", "DOB", "Gender", "Email", "State", "Address", "Skills"])

    # Fetch employee data
    employees = Employee.objects.all()
    for employee in employees:
        ws1.append(
            [
                employee.id,
                employee.name,
                employee.date_of_birth,
                employee.get_gender_display(),
                employee.email,
                employee.get_state_display(),
                employee.address,
                employee.skills,
            ]
        )
    # Adjust column widths and apply styles
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        adjusted_width = max_length + 2
        ws1.column_dimensions[col[0].column_letter].width = adjusted_width

    for row in ws1.iter_rows(min_col=3, max_col=3, min_row=2):
        for cell in row:
            if isinstance(cell.value, (datetime.date, datetime.datetime)):
                cell.number_format = "yyyy-mm-dd"  # Adjust format as needed

    # Prepare HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}"

    # Save the workbook to the response
    wb.save(response)
    return response


# working fine
def export_educations_to_excel(request):
    # Get file name from request (or set default value)
    file_name = request.GET.get("file_name", "educations.xlsx")

    # Create a workbook and sheets
    wb = Workbook()
    ws = wb.active
    ws.title = "Educations"

    # Define date style if necessary (for other use cases, not used here)
    # date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")

    # Add headers to Educations sheet
    ws.append(["Employee Name", "Education Type", "Description", "Graduation Year"])

    # Fetch education data
    educations = Education.objects.select_related("employee").all()
    for education in educations:
        ws.append(
            [
                education.employee.name,  # Use employee's name instead of ID
                education.edu_type,
                education.description,
                education.grad_year,
            ]
        )

    # Prepare HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={file_name}"

    # Save the workbook to the response
    wb.save(response)
    return response


# working allright , however i want some success msg and error msg if there is one
def import_employees_from_excel(request):
    if request.method == "POST":
        # Check if the request contains a file
        if "file" not in request.FILES:
            """ i don't see error msg in template """
            return render(request, "allEmployee.html", {"error": "No file selected"})

        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            return render(request, "allEmployee.html", {"error": "Invalid file format"})

        # Read the file using pandas
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return render(
                request, "allEmployee.html", {"error": f"Error reading file: {e}"}
            )

        # Process the data
        for _, row in df.iterrows():
            Employee.objects.create(
                # id=row['ID'],
                name=row["Name"],
                date_of_birth=row["DOB"],
                gender=row["Gender"],
                email=row["Email"],
                state=row["State"],
                address=row["Address"],
                skills=row["Skills"],
            )
        messages.success(request, "Employee imported successfully")
        # Redirect to a success page or display a success message
        return redirect(reverse("all_employee"))

    return render(request, "allEmployee.html")


def import_educations_from_excel(request):
    if request.method == "POST":
        if "file" not in request.FILES:
            messages.error(request, "No file selected")
            return redirect("all_employee")

        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "Invalid file format")
            return redirect("all_employee")

        try:
            df = pd.read_excel(file)
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("all_employee")

        for _, row in df.iterrows():
            Education.objects.create(
                employee=Employee.objects.get(name=row["Employee Name"]),
                edu_type=row["Education Type"],
                description=row["Description"],
                grad_year=row["Graduation Year"],
            )

        messages.success(request, "Educations imported successfully")
        return redirect(reverse("all_employee"))

    return render(request, "allEmployee.html")
